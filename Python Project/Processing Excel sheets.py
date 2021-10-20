import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook(filename="/home/pradumna/Desktop/Python/Python Project/transactions.xlsx")
sheet = wb['Sheet1']
# cell =sheet1['a1']
cell = sheet.cell(1,2)
print(cell.value)

# print(sheet1.max_row)

# Setting up the column header

# sheet.cell(1,4).value = "New Corrected Price"

for row in range(2,sheet.max_row +1):
    cell = sheet.cell(row , 3)
    corrected_price = cell.value*0.9
    correcetd_price_cell =sheet.cell(row,4)
    correcetd_price_cell.value = "$"+str(corrected_price)

values = Reference(sheet,min_row=2, max_row = sheet.max_row,min_col=4,max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')


wb.save('test1.xlsx')



